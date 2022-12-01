import csv
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, numbers
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


class DictionariesData:
    """Класс хранит в себе словари, нужные для работы программы.
    """
    dic_naming = {'name': 'Название',
                  'employer_name': 'Компания',
                  'description': 'Описание',
                  'key_skills': 'Навыки',
                  'experience_id': 'Опыт работы',
                  'premium': 'Премиум-вакансия',
                  'salary_from': 'Нижняя граница вилки оклада',
                  'salary_to': 'Верхняя граница вилки оклада',
                  'salary_gross': 'Оклад указан до вычета налогов',
                  'salary_currency': 'Идентификатор валюты оклада',
                  'area_name': 'Название региона',
                  'published_at': 'Дата публикации вакансии',
                  'Оклад': 'Оклад',
                  'False': 'Нет',
                  'True': 'Да',
                  'FALSE': 'Нет',
                  'TRUE': 'Да'
                  }


    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

class InputConect:
    """Класс отвечает за обработку параметров вводимых пользователем, а также за печать статистики на экран.

        Attributes:
            file_name (str): Название файла
            filter_param (str): Название профессии
    """
    def __init__(self):
        """Инициализирует объект InputConect.

            Args:
                file_name (str): Название файла
                filter_param (str): Название профессии
        """
        self.file_name = input('Введите название файла: ')
        self.filter_param = input('Введите название профессии: ')

    @staticmethod
    def print_data(dic_vacancies, filter_param):
        """Печатает статистику и вызывает методы для формирования графиков и отчетов.

            Args:
                dic_vacancies (list): Список вакансий
                filter_param (str): Название профессии
        """
        years = []
        for vacancy in dic_vacancies:
            years.append(int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')))
        years.sort()
        years = list(range(min(years), max(years) + 1))

        salary_by_years = {year: [] for year in years}
        vacs_by_years = {year: 0 for year in years}
        vac_salary_by_years = {year: [] for year in years}
        vac_counts_by_years = {year: 0 for year in years}

        for vacancy in dic_vacancies:
            year = int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
            salary_by_years[year].append(vacancy.salary.salary_to_rub)
            vacs_by_years[year] += 1
            if filter_param in vacancy.name:
                vac_salary_by_years[year].append(vacancy.salary.salary_to_rub)
                vac_counts_by_years[year] += 1
        salary_by_years = {key: int(sum(value)/ len(value)) if len(value) != 0 else 0 for key, value in salary_by_years.items()}
        vac_salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in vac_salary_by_years.items()}

        area_dic = {}
        for vacancy in dic_vacancies:
            if vacancy.area_name in area_dic:
                area_dic[vacancy.area_name].append(vacancy.salary.salary_to_rub)
            else:
                area_dic[vacancy.area_name] = [vacancy.salary.salary_to_rub]

        area_list = area_dic.items()
        area_list = [x for x in area_list if len(x[1]) / len(dic_vacancies) > 0.01]
        area_list.sort(key=lambda item: sum(item[1]) / len(item[1]), reverse=True)
        salary_by_cities = {item[0]: int(sum(item[1]) / len(item[1])) for item in area_list[0: min(len(area_list), 10)]}

        vacs_dic = {}
        for vacancy in dic_vacancies:
            if vacancy.area_name in vacs_dic:
                vacs_dic[vacancy.area_name] += 1
            else:
                vacs_dic[vacancy.area_name] = 1

        vacs_count = {x: round(y / len(dic_vacancies), 4) for x,y in vacs_dic.items()}
        vacs_count = {x: val for x, val in vacs_count.items() if val >= 0.01}
        vacs_by_cities = dict(sorted(vacs_count.items(), key=lambda item: item[1], reverse=True))
        others_percentage = sum(dict(list(vacs_by_cities.items())[11:]).values())
        vacs_by_cities = dict(list(vacs_by_cities.items())[:10])

        print('Динамика уровня зарплат по годам:', salary_by_years)
        print('Динамика количества вакансий по годам:', vacs_by_years)
        print('Динамика уровня зарплат по годам для выбранной профессии:', vac_salary_by_years)
        print('Динамика количества вакансий по годам для выбранной профессии:', vac_counts_by_years)
        print('Уровень зарплат по городам (в порядке убывания):', salary_by_cities)
        print('Доля вакансий по городам (в порядке убывания):', vacs_by_cities)

        report = Report(salary_by_years, vacs_by_years, vac_salary_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities, others_percentage, filter_param)
        Report.generate_excel(report)
        Report.generate_image(report)
        Report.generate_pdf(report)

class Report:
    """Класс отвечает за формирование графиков и отчетов.

    Attributes:
        salary_by_years (dict): Динамика уровня зарплат по годам
        vacs_by_years (dict): Динамика количества вакансий по годам
        vac_salary_by_years (dict): Динамика уровня зарплат по годам для выбранной профессии
        vac_counts_by_years (dict): Динамика количества вакансий по годам для выбранной профессии
        salary_by_cities (dict): Уровень зарплат по городам (в порядке убывания)
        vacs_by_cities (dict): Доля вакансий по городам (в порядке убывания)
        others_percentage (float): Доля вакансий по городам не входящих в Топ-10
        filter_param (str): Название профессии
    """
    def __init__(self, salary_by_years, vacs_by_years, vac_salary_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities, others_percentage, filter_param):
        """Инициализирует объект Report.

        Args:
            salary_by_years (dict): Динамика уровня зарплат по годам
            vacs_by_years (dict): Динамика количества вакансий по годам
            vac_salary_by_years (dict): Динамика уровня зарплат по годам для выбранной профессии
            vac_counts_by_years (dict): Динамика количества вакансий по годам для выбранной профессии
            salary_by_cities (dict): Уровень зарплат по городам (в порядке убывания)
            vacs_by_cities (dict): Доля вакансий по городам (в порядке убывания)
            others_percentage (float): Доля вакансий по городам не входящих в Топ-10
            filter_param (str): Название профессии
        """
        self.salary_by_years = salary_by_years
        self.vacs_by_years = vacs_by_years
        self.vac_salary_by_years = vac_salary_by_years
        self.vac_counts_by_years = vac_counts_by_years
        self.salary_by_cities = salary_by_cities
        self.vacs_by_cities = vacs_by_cities
        self.others_percentage = others_percentage
        self.filter_param = filter_param

    @staticmethod
    def as_text(value):
        """Парсит value в строку.

        Args:
            value (any): Входное значение, которое нужно конвертировать

        Returns:
            str: value конвертированное в строку.
        """
        if value is None:
            return ''
        return str(value)

    @staticmethod
    def generate_excel(report):
        """Генерирует excel файл с вакансиями.

           Args:
                report (Report): Объект класса Report
        """
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = 'Статистика по годам'
        sheet2 = wb.create_sheet('Статистика по городам')
        heads1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {report.filter_param}', 'Количество вакансий',
                  f'Количество вакансий - {report.filter_param}']
        heads2 = ['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий']
        for i, head in enumerate(heads1):
            sheet1.cell(1, i + 1, head).font = Font(bold=True)

        for year, value in report.salary_by_years.items():
            sheet1.append([year, value, report.vac_salary_by_years[year], report.vacs_by_years[year],
                           report.vac_counts_by_years[year]])

        for column_cells in sheet1.columns:
            lenght = max(len(Report.as_text(cell.value)) for cell in column_cells)
            sheet1.column_dimensions[column_cells[0].column_letter].width = lenght + 2

        thin = Side(border_style='thin', color='000000')
        for column in sheet1.columns:
            for cell in column:
                cell.border = Border(thin, thin, thin, thin)

        for i, head in enumerate(heads2):
            if i + 1 != 3:
                sheet2.cell(1, i + 1, head).font = Font(bold=True)

        values_count_salary = 10 if len(report.salary_by_cities) > 10 else len(report.salary_by_cities)
        values_count_vacs = 10 if len(report.vacs_by_cities) > 10 else len(report.vacs_by_cities)

        data = []
        for i, city in enumerate(report.salary_by_cities.items()):
            if i < values_count_salary:
                data.append(list(city) + [''])
            else:
                break

        for i, city in enumerate(report.vacs_by_cities.items()):
            if i < values_count_vacs:
                data[i] += list(city)

        for i, value in enumerate(data):
            sheet2.append(value)
            sheet2[f'E{i + 2}'].number_format = numbers.FORMAT_PERCENTAGE_00

        for column_cells in sheet2.columns:
            lenght = max(len(Report.as_text(cell.value)) for cell in column_cells)
            sheet2.column_dimensions[column_cells[0].column_letter].width = lenght + 2

        for row in sheet2.rows:
            for cell in row:
                if cell.value != None and cell.value != '':
                    cell.border = Border(thin, thin, thin, thin)

        wb.save('report.xlsx')


    @staticmethod
    def generate_image(report):
        """Генерирует png файл с графиками.

           Args:
                report (Report): Объект класса Report
        """
        width = 0.4
        x_nums = np.arange(len(report.salary_by_years.keys()))
        x_list1 = x_nums - width/2
        x_list2 = x_nums + width/2

        fig = plt.figure(figsize=(8, 6))
        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(x_list1, report.salary_by_years.values(), width, label='средняя з/п')
        ax.bar(x_list2, report.vac_salary_by_years.values(), width, label=f'з/п {(report.filter_param).lower()}')
        ax.set_xticks(x_nums, report.salary_by_years.keys(), rotation='vertical')
        ax.legend(fontsize=8, loc='upper left')
        ax.tick_params(axis='both', labelsize= 8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансий по годам")
        ax.bar(x_list1, report.vacs_by_years.values(), width, label='Количество вакансий')
        ax.bar(x_list2, report.vac_counts_by_years.values(), width, label=f'Количество вакансий \n{(report.filter_param).lower()}')
        ax.set_xticks(x_nums, report.vacs_by_years.keys(), rotation='vertical')
        ax.legend(fontsize=8, loc='upper left')
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        y = list(reversed(report.salary_by_cities.keys()))
        y = [x.replace(' ', '\n').replace('-','-\n') for x in y]
        x = list(reversed(report.salary_by_cities.values()))
        plt.barh(y, x)
        ax.tick_params(axis='y', labelsize=6)
        ax.tick_params(axis='x', labelsize=8)
        ax.grid(True, axis='x')

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансий по городам")
        cities = list(report.vacs_by_cities.keys())
        percents = list(report.vacs_by_cities.values())
        if report.others_percentage != 0:
            cities.insert(0,'Другие')
            percents.insert(0, 1-sum(percents))
        plt.pie(percents, labels=cities, textprops={'fontsize': 6}, startangle =90)

        plt.tight_layout()
        plt.savefig('graph.png')

    @staticmethod
    def generate_pdf(report):
        """Генерирует pdf файл из png и excel файлов.

           Args:
                report (Report): Объект класса Report
        """
        vacancy_name = report.filter_param
        image_file = 'graph.png'

        options = {
            "enable-local-file-access": None
        }

        heads1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {vacancy_name}', 'Количество вакансий',
                  f'Количество вакансий - {vacancy_name}']
        heads2 = ['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий']

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        report.vacs_by_cities = {key: (str(round(float(value) * 100, 3))).replace('.', ',')+'%' for key, value in
                              report.vacs_by_cities.items()}

        pdf_template = template.render({'vacancy': vacancy_name, 'image_file': image_file,
                                        "vac_salary_by_years": report.vac_salary_by_years,
                                        "vacs_by_years": report.vacs_by_years,
                                        "vac_counts_by_years": report.vac_counts_by_years,
                                        "salary_by_years": report.salary_by_years,
                                        "salary_by_cities": report.salary_by_cities,
                                        "vacs_by_cities": report.vacs_by_cities,
                                        "heads1": heads1,
                                        "heads2": heads2})
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

class DataSet:
    """Класс отвечает за чтение и подготовку данных из CSV-файла.

        Attributes:
            file_name (str): Название файла
            vacancies_objects (list): Список вакансий
    """
    def __init__(self, file_name):
        """Инициализирует объект DataSet.

        Args:
            file_name (str): Название файла
            vacancies_objects (list): Список вакансий
        """
        self.file_name = file_name
        self.vacancies_objects = DataSet.prepare_data(file_name)

    @staticmethod
    def Prepare(text):
        """Подготавливает строку: убирает лишние пробелы, html-теги, заменяет \n на ; для дальнейшего split'а.

        Args:
            text (str): Строка, которую нужно очистить

        Returns:
            str: Очищенная строка
        """
        text = re.sub(r"<[^>]+>", '', text)
        text = "; ".join(text.split('\n'))
        text = ' '.join(text.split())
        return text

    @staticmethod
    def prepare_data(file_name):
        """Отбирает вакансии без пустых ячеек и составляет лист вакансий.

        Args:
            file_name (str): Название файла

        Returns:
            list: Лист, состоящий из вакансий.
        """
        columns, vacancies = DataSet.csv_reader(file_name)
        processed = [x for x in vacancies if len(x) == len(columns) and '' not in x]
        people_data = []
        for line in processed:
            vacancies_list = {}
            salary = []
            for i,item in enumerate(line):
                if columns[i] in ['salary_from', 'salary_to', 'salary_currency']:
                    salary.append(DataSet.Prepare(item))
                    if len(salary) == 3:
                        vacancies_list['salary'] = Salary(salary[0], salary[1], salary[2])
                else:
                    vacancies_list[columns[i]] = (DataSet.Prepare(item))
            people_data.append(Vacancy(vacancies_list['name'], vacancies_list['salary'], vacancies_list['area_name'], vacancies_list['published_at']))
        return people_data

    @staticmethod
    def csv_reader(file_name):
        """Считывает csv файл.

        Args:
            file_name (str): Название файла

        Returns:
            tuple: Кортеж, состоящий из названий колонок и всех вакансий.
        """
        csv_read = csv.reader(open(file_name, encoding="utf_8_sig"))
        list_data = [x for x in csv_read]
        if len(list_data) == 0:
            print("Пустой файл")
            exit()
        columns = list_data[0]
        vacancies = list_data[1:]
        return columns, vacancies

class Salary:
    """Класс для представления зарплаты.

    Attributes:
        salary_from (str): Нижняя граница вилки оклада
        salary_to (str): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
        salary_to_rub (int): Средняя зарплата в рублях
    """
    def __init__(self, salary_from, salary_to, salary_currency):
        """Инициализирует объект Salary.

        Args:
            salary_from (str or int or float): Нижняя граница вилки оклада
            salary_to (str or int or float): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
            salary_to_rub (int): Средняя зарплата в рублях
        """

        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_to_rub = Salary.currency_to_rub(salary_from, salary_to, salary_currency)

    @staticmethod
    def currency_to_rub(salary_from, salary_to, salary_currency):
        """Вычисляет среднюю зарплату из вилки и переводит в рубли, при помощи словаря - currency_to_rub.

        Args:
            salary_from (str): Нижняя вилка оклада
            salary_to (str): Верхняя вилка оклада
            salary_currency (str): Валюта оклада

        Returns:
            float: Средняя зарплата в рублях
        """
        return (float(salary_from) + float(salary_to)) / 2 * DictionariesData.currency_to_rub[salary_currency]


class Vacancy:
    """Класс устанавливает все основные поля вакансии.

        Attributes:
            name (str): Название вакансии
            salary (Salary): Комбинированная информация о зарплате
            area_name (str): Название региона
            published_at (str): Дата публикации вакансии
    """
    def __init__(self, name, salary, area_name, published_at):
        """Инициализирует объект Vacancy.

                Args:
                    name (str): Название вакансии
                    salary (Salary): Комбинированная информация о зарплате
                    area_name (str): Название региона
                    published_at (str): Дата публикации вакансии
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at

def get_table():
    """Используется в main.py. Формирует pdf файл.
    """
    parameters = InputConect()
    dataset = DataSet(parameters.file_name)
    InputConect.print_data(dataset.vacancies_objects, parameters.filter_param)